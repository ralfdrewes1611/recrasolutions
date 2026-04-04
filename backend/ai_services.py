"""
RECRA AI Services — Excel import, Website scraper, Floor plan analysis, Quote text generation
"""
import os
import json
import uuid
import logging
import io
import csv
import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent

logger = logging.getLogger("ai_services")
ai_router = APIRouter(prefix="/api/ai")


def get_chat(system_message: str) -> LlmChat:
    llm_key = os.environ.get('EMERGENT_LLM_KEY')
    if not llm_key:
        raise HTTPException(status_code=500, detail="AI niet geconfigureerd (EMERGENT_LLM_KEY ontbreekt)")
    return LlmChat(
        api_key=llm_key,
        session_id=f"recra-ai-{uuid.uuid4()}",
        system_message=system_message
    ).with_model("openai", "gpt-5.2")


def parse_json_response(response: str) -> Any:
    """Extract JSON from AI response, handles markdown code blocks."""
    text = response.strip()
    if '```json' in text:
        text = text.split('```json')[1].split('```')[0].strip()
    elif '```' in text:
        text = text.split('```')[1].split('```')[0].strip()
    
    # Try array first, then object
    for start_char, end_char in [('[', ']'), ('{', '}')]:
        start = text.find(start_char)
        end = text.rfind(end_char)
        if start >= 0 and end > start:
            try:
                return json.loads(text[start:end + 1])
            except json.JSONDecodeError:
                continue
    raise ValueError(f"Could not parse JSON from: {text[:200]}")


# ─── 1. EXCEL / CSV IMPORT ─────────────────────────────────────

class ImportedProduct(BaseModel):
    name: str
    category: str = "overig"
    description: str = ""
    price_purchase: float = 0
    price_lease_monthly: float = 0
    installation_cost: float = 0
    maintenance_yearly: float = 0
    dimensions_width: float = 1
    dimensions_height: float = 1
    confidence: float = 1.0


class ImportPreview(BaseModel):
    products: List[ImportedProduct]
    column_mapping: Dict[str, str]
    raw_headers: List[str]
    total_rows: int


@ai_router.post("/import-products", response_model=ImportPreview)
async def import_products_file(file: UploadFile = File(...)):
    """Upload Excel/CSV → AI herkent kolommen → geeft preview van producten."""
    
    content = await file.read()
    filename = file.filename.lower()
    
    # Read file into rows
    headers = []
    rows = []
    
    if filename.endswith(('.xlsx', '.xls')):
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise HTTPException(status_code=400, detail="Leeg bestand")
        headers = [str(h or '').strip() for h in all_rows[0]]
        rows = [[str(c or '') for c in row] for row in all_rows[1:] if any(c for c in row)]
        wb.close()
    elif filename.endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.reader(io.StringIO(text), delimiter=';')
        all_rows = list(reader)
        if not all_rows:
            # Try comma delimiter
            reader = csv.reader(io.StringIO(text), delimiter=',')
            all_rows = list(reader)
        if not all_rows:
            raise HTTPException(status_code=400, detail="Leeg bestand")
        headers = [h.strip() for h in all_rows[0]]
        rows = [row for row in all_rows[1:] if any(c.strip() for c in row)]
    else:
        raise HTTPException(status_code=400, detail="Alleen .xlsx, .xls of .csv bestanden worden ondersteund")
    
    if not rows:
        raise HTTPException(status_code=400, detail="Geen data gevonden in het bestand")
    
    # Send headers + sample data to AI for column mapping
    sample_data = rows[:5]
    sample_text = f"Headers: {headers}\n\nVoorbeeld data (eerste {len(sample_data)} rijen):\n"
    for i, row in enumerate(sample_data):
        sample_text += f"Rij {i+1}: {dict(zip(headers, row))}\n"
    
    chat = get_chat("""Je bent een expert in data-import voor recreatieparken. 
    Analyseer de kolomnamen en data, en map ze naar het RECRA productschema.
    
    Mogelijke velden: name, category, description, price_purchase, price_lease_monthly, 
    installation_cost, maintenance_yearly, dimensions_width, dimensions_height
    
    Mogelijke categorieën: sanitair, slagboom, camera, wifi, verlichting, betaalsysteem, 
    toegangscontrole, douchelezer, energie, overig
    
    Geef je antwoord als JSON:
    {
        "column_mapping": {"originele_kolomnaam": "recra_veld", ...},
        "category_guess": "meest waarschijnlijke categorie voor deze producten"
    }
    
    Map alleen kolommen die je zeker herkent. Onbekende kolommen negeer je.""")
    
    response = await chat.send_message(UserMessage(text=sample_text))
    mapping_result = parse_json_response(response)
    column_mapping = mapping_result.get("column_mapping", {})
    default_category = mapping_result.get("category_guess", "overig")
    
    # Apply mapping to all rows
    products = []
    for row in rows:
        row_dict = dict(zip(headers, row))
        product_data = {"category": default_category}
        
        for orig_col, target_field in column_mapping.items():
            value = row_dict.get(orig_col, "")
            if not value:
                continue
            
            if target_field in ('price_purchase', 'price_lease_monthly', 'installation_cost', 
                                'maintenance_yearly', 'dimensions_width', 'dimensions_height'):
                # Parse number (handles €, comma, etc.)
                cleaned = value.replace('€', '').replace('.', '').replace(',', '.').strip()
                try:
                    product_data[target_field] = float(cleaned)
                except ValueError:
                    pass
            else:
                product_data[target_field] = value.strip()
        
        if product_data.get('name'):
            products.append(ImportedProduct(**{
                **{"name": "", "category": default_category},
                **product_data,
                "confidence": 0.9 if 'name' in column_mapping.values() else 0.5,
            }))
    
    return ImportPreview(
        products=products,
        column_mapping=column_mapping,
        raw_headers=headers,
        total_rows=len(rows),
    )


class ConfirmImportRequest(BaseModel):
    products: List[ImportedProduct]


@ai_router.post("/import-products/confirm")
async def confirm_import_products(request: ConfirmImportRequest):
    """Bevestig en sla geïmporteerde producten op in de database."""
    from motor.motor_asyncio import AsyncIOMotorClient
    mongo_url = os.environ['MONGO_URL']
    client = AsyncIOMotorClient(mongo_url)
    db_local = client[os.environ['DB_NAME']]
    
    saved = 0
    for product in request.products:
        doc = {
            "id": str(uuid.uuid4()),
            "name": product.name,
            "category": product.category,
            "description": product.description,
            "price_purchase": product.price_purchase,
            "price_lease_monthly": product.price_lease_monthly,
            "installation_cost": product.installation_cost,
            "maintenance_yearly": product.maintenance_yearly,
            "dimensions": {"width": product.dimensions_width, "height": product.dimensions_height},
            "icon": "package",
            "color": "#70C26C",
            "created_at": __import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat(),
        }
        await db_local.products.insert_one(doc)
        saved += 1
    
    client.close()
    return {"saved": saved, "message": f"{saved} producten geïmporteerd"}


# ─── 2. WEBSITE SCRAPER ────────────────────────────────────────

class ScrapeRequest(BaseModel):
    url: str


class ScrapedProduct(BaseModel):
    name: str = ""
    description: str = ""
    price: float = 0
    dimensions_width: float = 1
    dimensions_height: float = 1
    category: str = "overig"
    source_url: str = ""
    image_url: str = ""


class ScrapeResult(BaseModel):
    products: List[ScrapedProduct]
    page_title: str = ""
    source_url: str = ""


@ai_router.post("/scrape-products", response_model=ScrapeResult)
async def scrape_products(request: ScrapeRequest):
    """Scrape een product-pagina en gebruik AI om productinfo te extraheren."""
    
    url = request.url
    if not url.startswith('http'):
        url = 'https://' + url
    
    # Fetch the page
    async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
        try:
            resp = await client.get(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            })
            resp.raise_for_status()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Kon pagina niet laden: {str(e)}")
    
    # Parse HTML
    soup = BeautifulSoup(resp.text, 'html.parser')
    
    # Remove scripts, styles
    for tag in soup(['script', 'style', 'nav', 'footer', 'header']):
        tag.decompose()
    
    page_title = soup.title.string if soup.title else ""
    
    # Get text content (limit to avoid token overflow)
    text_content = soup.get_text(separator='\n', strip=True)[:6000]
    
    # Get images
    images = []
    for img in soup.find_all('img', src=True)[:10]:
        src = img['src']
        if src.startswith('//'):
            src = 'https:' + src
        elif src.startswith('/'):
            from urllib.parse import urlparse
            parsed = urlparse(url)
            src = f"{parsed.scheme}://{parsed.netloc}{src}"
        alt = img.get('alt', '')
        images.append(f"{alt}: {src}")
    
    # Send to AI for extraction
    chat = get_chat("""Je bent een expert in het extraheren van productinformatie uit webpagina's voor recreatieparken.
    
    Analyseer de tekst en afbeeldingen van de pagina en extraheer alle producten die je vindt.
    
    Voor elk product, geef:
    - name: productnaam
    - description: korte beschrijving
    - price: aankoopprijs in euro (0 als niet gevonden)
    - dimensions_width: breedte in meters (schat als niet gevonden)
    - dimensions_height: diepte/lengte in meters (schat als niet gevonden)
    - category: kies uit sanitair, slagboom, camera, wifi, verlichting, betaalsysteem, toegangscontrole, douchelezer, energie, overig
    - image_url: URL van productafbeelding (als gevonden)
    
    Geef je antwoord als JSON array:
    [{"name": "...", "description": "...", "price": 0, "dimensions_width": 1, "dimensions_height": 1, "category": "overig", "image_url": ""}]
    
    Geef alleen echte producten terug, geen navigatie-items of categorieën.""")
    
    prompt = f"Pagina URL: {url}\nTitel: {page_title}\n\nTekst:\n{text_content}\n\nAfbeeldingen:\n" + "\n".join(images)
    
    response = await chat.send_message(UserMessage(text=prompt))
    
    try:
        products_data = parse_json_response(response)
        if not isinstance(products_data, list):
            products_data = [products_data]
        
        products = [ScrapedProduct(
            **{k: v for k, v in p.items() if k in ScrapedProduct.model_fields},
            source_url=url
        ) for p in products_data]
    except Exception as e:
        logger.error(f"Scrape parse error: {e}")
        products = []
    
    return ScrapeResult(
        products=products,
        page_title=page_title,
        source_url=url,
    )


# ─── 3. AI QUOTE TEXT GENERATION ───────────────────────────────

class QuoteTextRequest(BaseModel):
    project_name: str
    project_type: str
    num_spots: int
    products: List[Dict[str, Any]]
    total_investment: float
    lease_monthly: float


class QuoteTextResponse(BaseModel):
    intro: str
    body: str
    closing: str


@ai_router.post("/generate-quote-text", response_model=QuoteTextResponse)
async def generate_quote_text(request: QuoteTextRequest):
    """Genereer een professionele offertetekst in het Nederlands."""
    
    product_summary = "\n".join([
        f"- {p.get('name', '')}: {p.get('quantity', 1)}x (€{p.get('price', 0):,.0f})"
        for p in request.products
    ])
    
    chat = get_chat("""Je bent een professionele offerteschrijver voor RECRA Solutions, 
    specialist in recreatieparken en campings.
    
    Schrijf een overtuigende maar zakelijke offertetekst in het Nederlands.
    Gebruik een professionele toon, benoem de voordelen voor de klant.
    
    Geef je antwoord als JSON:
    {
        "intro": "Inleidende alinea (2-3 zinnen, persoonlijke aanspraak, verwijs naar het project)",
        "body": "Hoofdtekst (3-4 alinea's: samenvatting oplossing, technische highlights, service/SLA)",
        "closing": "Afsluiting (2 zinnen, call to action, contactinfo RECRA Solutions)"
    }""")
    
    prompt = f"""Schrijf een offerte voor:
Projectnaam: {request.project_name}
Type: {request.project_type}
Standplaatsen: {request.num_spots}

Producten:
{product_summary}

Totaal investering: €{request.total_investment:,.0f}
Operational Lease: €{request.lease_monthly:,.0f}/maand (60 maanden incl. SLA)"""
    
    response = await chat.send_message(UserMessage(text=prompt))
    
    try:
        result = parse_json_response(response)
        return QuoteTextResponse(**result)
    except Exception as e:
        logger.error(f"Quote text error: {e}")
        return QuoteTextResponse(
            intro=f"Geachte klant, graag presenteren wij u onze offerte voor {request.project_name}.",
            body=f"Op basis van uw wensen hebben wij een complete configuratie samengesteld voor {request.num_spots} standplaatsen. De totale investering bedraagt €{request.total_investment:,.0f}, of €{request.lease_monthly:,.0f} per maand via Operational Lease (60 maanden incl. SLA).",
            closing="Wij gaan graag met u in gesprek om deze offerte verder te bespreken. Neem contact op met RECRA Solutions."
        )


# ─── 4. ENHANCED FLOOR PLAN ANALYSIS ──────────────────────────

class FloorPlanAnalysisRequest(BaseModel):
    image_base64: str
    project_type: str = "camping"
    canvas_width: int = 1000
    canvas_height: int = 700


class DetectedZone(BaseModel):
    name: str
    type: str
    color: str
    points: List[Dict[str, float]]


class FloorPlanResult(BaseModel):
    zones: List[DetectedZone]
    estimated_spots: int = 0
    suggested_scale: float = 0.1
    suggestions: List[str] = []


@ai_router.post("/analyze-floorplan-smart", response_model=FloorPlanResult)
async def analyze_floorplan_smart(request: FloorPlanAnalysisRequest):
    """Analyseer plattegrond met Vision AI en genereer zones op het canvas."""
    
    chat = get_chat(f"""Je bent een expert in het analyseren van plattegronden voor recreatieparken.
    
    Analyseer de afbeelding en identificeer:
    1. Wegen/paden (type: toegangsweg, kleur: #9ca3af)
    2. Standplaatsen/kavels (type: standplaats, kleur: #70C26C)
    3. Sanitairgebouwen (type: sanitair, kleur: #0ea5e9)
    4. Speeltuinen/faciliteiten (type: faciliteit, kleur: #f59e0b)
    5. Water/vijvers (type: water, kleur: #3b82f6)
    6. Bomen/groen (type: groen, kleur: #166534)
    
    Het canvas is {request.canvas_width}x{request.canvas_height} pixels.
    Geef per zone de hoekpunten als pixels binnen het canvas.
    
    Geef je antwoord als JSON:
    {{
        "zones": [
            {{"name": "Naam", "type": "type", "color": "#kleur", "points": [{{"x": 0, "y": 0}}, ...]}}
        ],
        "estimated_spots": getal,
        "suggested_scale": meter_per_pixel (bijv 0.1 = 10m per 100px),
        "suggestions": ["tip 1", "tip 2"]
    }}
    
    Maak realistische zones die het terrein representeren.""")
    
    image_b64 = request.image_base64
    if ',' in image_b64:
        image_b64 = image_b64.split(',')[1]
    
    response = await chat.send_message(UserMessage(
        text=f"Analyseer deze plattegrond voor een {request.project_type}. Canvas: {request.canvas_width}x{request.canvas_height}px.",
        file_contents=[ImageContent(image_base64=image_b64)]
    ))
    
    try:
        result = parse_json_response(response)
        zones = [DetectedZone(**z) for z in result.get("zones", [])]
        return FloorPlanResult(
            zones=zones,
            estimated_spots=result.get("estimated_spots", 0),
            suggested_scale=result.get("suggested_scale", 0.1),
            suggestions=result.get("suggestions", []),
        )
    except Exception as e:
        logger.error(f"Floor plan analysis error: {e}")
        return FloorPlanResult(
            zones=[],
            estimated_spots=0,
            suggested_scale=0.1,
            suggestions=["Kon de plattegrond niet automatisch analyseren. Teken zones handmatig."],
        )
